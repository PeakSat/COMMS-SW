import openpyxl
import requests
import os

# Google Sheets details
spreadsheet_id = "12m4Sq4CMnLB9nn6INxKUdSNSIlTJes_uTbg19RE9-X0"

# Construct the export URL for .xlsx format
xlsx_export_url = f"https://docs.google.com/spreadsheets/d/{spreadsheet_id}/export?format=xlsx"

# Output file
output_file = "parameter_database.xlsx"

try:
    # Send GET request
    response = requests.get(xlsx_export_url)

    # Check if the request was successful
    if response.status_code == 200:
        # Write the content to a file
        with open(output_file, "wb") as file:
            file.write(response.content)
        print(f"Excel file downloaded successfully as '{output_file}'")
    else:
        print(f"Failed to download Excel file. HTTP Status Code: {response.status_code}")
except Exception as e:
    print(f"An error occurred: {e}")


# Subsystem acronyms and their corresponding number offsets
subsystem_config = {
    "OBDH": 0,
    "COMMS": 819,
    "PAY": 1638,
    "ADCS": 2457,
    "EPS": 3276,
}

# Type encoding dictionary for the last 4 bits
type_encoding = {
    "uint8_t": 0,
    "bool" : 0,
    "int8_t": 1,
    "uint16_t": 2,
    "int16_t": 3,
    "uint32_t": 4,
    "int32_t": 5,
    "uint64_t": 6,
    "int64_t": 7,
    "float": 8,
    "double": 9,
}

# Function to encode the ID
def encode_id(numeric_id, variable_type):
    if variable_type not in type_encoding:
        print(f"Error: Type '{variable_type}' not found in type encoding dictionary. Defaulting to 'uint64_t'.")
        variable_type = "uint64_t"
    type_code = type_encoding[variable_type]
    return (numeric_id << 4) | type_code

# Paths to the Excel file and output directories
excel_file = "parameter_database.xlsx"
src_folder = "Src"
inc_folder = "Inc"
output_cpp_file = os.path.join(src_folder, "PlatformParameters.cpp")
output_hhp_file = os.path.join(inc_folder, "PlatformParameters.hpp")

# Ensure the output directories exist
os.makedirs(src_folder, exist_ok=True)
os.makedirs(inc_folder, exist_ok=True)

# Set to track unique IDs
processed_ids = set()
parameter_types = {}  # Dictionary to store {parameter_name: variable_type}

# Load the workbook
workbook = openpyxl.load_workbook(excel_file)

# Lists to collect output lines
cpp_lines = []
hhp_lines = []

# Add the header lines for the .cpp file
cpp_lines.append('#include "COMMS_ECSS_Configuration.hpp"')
cpp_lines.append("\n#ifdef SERVICE_PARAMETER\n")
cpp_lines.append('#include "Services/ParameterService.hpp"')
cpp_lines.append('#include "PlatformParameters.hpp"')
cpp_lines.append("\nvoid ParameterService::initializeParameterMap() {")
cpp_lines.append("    parameters = {")

# Add the header lines for the .hpp file
hhp_lines.append("#pragma once")
hhp_lines.append("#pragma GCC diagnostic push")
hhp_lines.append('#pragma GCC diagnostic ignored "-Wpsabi" // Suppress: parameter passing for argument of type \'Time::DefaultCUC\' {aka \'TimeStamp<4, 0, 1, 10>\'} changed in GCC 7.1')
hhp_lines.append('#include "Helpers/Parameter.hpp"')

# Process each subsystem separately
namespace_blocks = {acronym: [] for acronym in subsystem_config.keys()}
valid_rows = []

# Collect all valid rows across all sheets
for sheet_name in workbook.sheetnames:
    sheet = workbook[sheet_name]
    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row):
        id_cell = row[0]  # First column
        variable_cell = row[4]  # Fifth column

        # Skip rows with no variable name
        if not variable_cell.value or not isinstance(variable_cell.value, str):
            continue

        # Check if the first column contains a valid ID
        if id_cell.value and isinstance(id_cell.value, str):
            # Add to valid rows
            valid_rows.append(row)

# Process all valid rows
for idx, row in enumerate(valid_rows):
    id_cell = row[0]  # First column
    type_cell = row[2]  # Third column
    variable_cell = row[4]  # Fifth column
    enum_items_cell = row[6]  # Seventh column
    value_cell = row[7]  # Eighth column

    # Identify the subsystem acronym
    for acronym, offset in subsystem_config.items():
        if id_cell.value.startswith(f"{acronym}-"):
            # Extract the numeric part of the ID
            numeric_id = id_cell.value[len(acronym) + 1:]  # Remove acronym and '-'

            try:
                numeric_id = int(numeric_id) + offset
            except ValueError:
                print(f"Skipping invalid numeric ID: {id_cell.value}")
                continue

            # Get variable name and type
            variable_name = variable_cell.value.strip()
            variable_type = type_cell.value.strip() if type_cell.value else "int"
            enum_items = enum_items_cell.value.strip() if enum_items_cell.value else ""

            # Handle "_enum" type
            if variable_type.endswith("_enum"):
                base_param = variable_type[:-5]  # Remove "_enum"
                if base_param in parameter_types:
                    variable_type = parameter_types[base_param]
                else:
                    print(f"Error: Parameter '{base_param}' not found for '{variable_name}_enum'. Defaulting to 'uint32_t'.")
                    variable_type = "uint32_t"

            # Store the type in the parameter_types dictionary
            parameter_types[variable_name] = variable_type

            # Encode the ID with the new encoding rule
            encoded_id = encode_id(numeric_id, variable_type)

            # Skip duplicates
            if encoded_id in processed_ids:
                continue
            processed_ids.add(encoded_id)

            # Handle float values to remove .0 for whole numbers
            if value_cell.value:
                if isinstance(value_cell.value, float) and value_cell.value.is_integer():
                    param_value = str(int(value_cell.value))  # Convert to int if it's a whole number
                else:
                    param_value = str(value_cell.value).strip()  # Convert to string for other cases
            else:
                param_value = "0"

            # Add to the corresponding namespace block
            block_lines = namespace_blocks[acronym]
            block_lines.append(f"        {variable_name}ID = {encoded_id}")

            variable_type = type_cell.value.strip() if type_cell.value else "int"
            # Enum definitions (if type is "enum")
            if (variable_type in {"uint8_t", "uint16_t", "uint32_t", "uint64_t",
                                  "int8_t", "int16_t", "int32_t", "int64_t"} and row[6].value and str(row[6].value).strip()):  # 7th column (index 6)
                enum_lines = (
                    f"    enum {variable_name}_enum : {type_cell.value.strip()} {{\n        {enum_items}\n    }};"
                )
                block_lines.append(enum_lines)
                param_line = f"    inline Parameter<{variable_name}_enum> {variable_name}({param_value});"
            # Parameter initializations
            elif variable_type == "enum":
                param_line = f"    inline Parameter<{variable_name}_enum> {variable_name}({param_value});"
            else:
                param_line = f"    inline Parameter<{variable_type}> {variable_name}({param_value});"
            block_lines.append(param_line)

            # Add to .cpp file
            is_last_row = idx == len(valid_rows) - 1
            if is_last_row:
                cpp_lines.append(
                    f'        {{{acronym}Parameters::{variable_name}ID, {acronym}Parameters::{variable_name}}}'
                )
            else:
                cpp_lines.append(
                    f'        {{{acronym}Parameters::{variable_name}ID, {acronym}Parameters::{variable_name}}},'
                )

            break

# Finalize the .cpp file
cpp_lines.append("    };")
cpp_lines.append("}")
cpp_lines.append("#endif")

# Build the .hhp file
for acronym, block_lines in namespace_blocks.items():
    if block_lines:
        hhp_lines.append(f"namespace {acronym}Parameters {{")
        hhp_lines.append("    enum ParameterID : uint16_t {")
        hhp_lines.append(",\n".join(line for line in block_lines if "ID =" in line))
        hhp_lines.append("    };")
        hhp_lines.extend(
            line for line in block_lines if not line.startswith("        ")
        )
        hhp_lines.append("}")

# Add footer to the .hhp file
hhp_lines.append("#pragma GCC diagnostic pop")

# Write the .cpp file
with open(output_cpp_file, "w") as cpp_file:
    cpp_file.write("\n".join(cpp_lines))

# Write the .hhp file
with open(output_hhp_file, "w") as hhp_file:
    hhp_file.write("\n".join(hhp_lines))

print(f"Processing complete.")
print(f"CPP file saved as '{output_cpp_file}'.")
print(f"HHP file saved as '{output_hhp_file}'.")